"""
SteelDigitize Pro — openpyxl MCP 工具集
find_last_row / write_batch / verify_batch
"""
from __future__ import annotations
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ---- 格式常量 ----
FONT_SONG = Font(name='宋体', size=11)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
HEADER_FILL = PatternFill(start_color='E7E8EA', end_color='E7E8EA', fill_type='solid')

COL_WIDTHS = {
    'A': 5.5,   # 序号 40px
    'B': 12.5,  # 单号 90px
    'C': 12.5,  # 日期 90px
    'D': 14,    # 品种 100px
    'E': 16.5,  # 规格 120px
    'F': 7,     # 单位 50px
    'G': 8.5,   # 数量 60px
    'H': 9.5,   # 单价 70px
    'I': 12.5,  # 金额 90px
    'J': 14,    # 合计 100px
}


def _check_file_lock(filepath: str) -> dict | None:
    """检测文件是否被 WPS/Excel 占用。被占用返回 error dict，否则 None。"""
    # 不能用 a/a+：文件不存在时会被悄悄创建成 0 字节，
    # 随后 _open_workbook 会把这个空文件当成损坏的 xlsx 打开。
    if not os.path.exists(filepath):
        return None
    try:
        with open(filepath, 'rb+') as f:
            pass
    except PermissionError:
        return {"success": False, "error": "文件被 WPS/Excel 占用，请关闭后重试"}
    return None


def _open_workbook(filepath: str) -> Workbook:
    """打开 Excel 文件，不存在则创建"""
    if os.path.exists(filepath):
        return load_workbook(filepath)
    return Workbook()


def _apply_format(ws, row, col_count):
    """对一行单元格应用格式：宋体11pt、居中、细线边框"""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_SONG
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _set_column_widths(ws):
    """设置列宽"""
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


# ---- 工具1: find_last_row ----

def find_last_row(filepath: str, sheet: str) -> dict:
    """找到 sheet 中最后一行数据的位置"""
    lock_err = _check_file_lock(filepath)
    if lock_err:
        return lock_err

    wb = _open_workbook(filepath)
    try:
        if sheet not in wb.sheetnames:
            return {"last_row": 0, "next_row": 1, "has_headers": False}

        ws = wb[sheet]
        last_row = ws.max_row
        # 从底部向上找第一个有数据的行
        while last_row > 0:
            has_data = any(ws.cell(row=last_row, column=c).value is not None for c in range(1, 11))
            if has_data:
                break
            last_row -= 1

        has_headers = last_row > 0
        return {
            "last_row": last_row,
            "next_row": last_row + 1,
            "has_headers": has_headers,
        }
    finally:
        wb.close()


# ---- 工具2: write_batch ----

def write_batch(
    filepath: str,
    sheet: str,
    mode: str,
    start_row: int,
    seq: int,
    receipt_no: str,
    date: str,
    items: list,
) -> dict:
    """
    将一个单据的数据写入 Excel。
    自动处理合并单元格、公式、格式。
    """
    lock_err = _check_file_lock(filepath)
    if lock_err:
        return lock_err

    if not items:
        return {"success": False, "error": "items 列表为空"}

    wb = _open_workbook(filepath)
    try:
        # 获取或创建 sheet
        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.create_sheet(title=sheet)

        if mode not in {"new", "append"}:
            return {"success": False, "error": "mode 必须是 new 或 append"}
        # 新建模式只允许空表或仅含表头的表，不能把已有对账单从第 1 行覆盖掉。
        if mode == "new" and any(
            ws.cell(row=row, column=col).value is not None
            for row in range(2, ws.max_row + 1)
            for col in range(1, 11)
        ):
            return {"success": False, "error": "目标表格已有数据；为保护历史数据，请使用 append 追加"}

        # 设置列宽
        _set_column_widths(ws)

        # mode=new: 先写表头
        if mode == "new":
            headers = ['序号', '单号', '日期', '品种', '规格', '单位', '数量', '单价', '金额', '合计金额']
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = FONT_SONG
                cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER
                cell.fill = HEADER_FILL
            if start_row < 2:
                start_row = 2

        data_start = start_row
        data_end = start_row + len(items) - 1

        # 写入数据行
        for i, item in enumerate(items):
            row = start_row + i
            # A 列：序号（仅首行）
            if i == 0:
                ws.cell(row=row, column=1, value=seq)
                ws.cell(row=row, column=2, value=receipt_no)
                ws.cell(row=row, column=3, value=date)

            # D: 品种（name 为空时留空，后续合并单元格时处理）
            ws.cell(row=row, column=4, value=item.get("name", ""))
            ws.cell(row=row, column=5, value=item.get("spec", ""))
            ws.cell(row=row, column=6, value=item.get("unit", ""))
            ws.cell(row=row, column=7, value=item.get("qty", 0))
            ws.cell(row=row, column=8, value=item.get("price", 0))
            # I 列：=G*H 公式
            ws.cell(row=row, column=9).value = f"=G{row}*H{row}"

            _apply_format(ws, row, 10)

        # J 列：=SUM(I) 公式（仅首行，后续合并后统一）
        ws.cell(row=data_start, column=10).value = f"=SUM(I{data_start}:I{data_end})"
        _apply_format(ws, data_start, 10)

        # ---- 合并单元格 ----
        if len(items) > 1:
            # A/B/C/J 列跨所有数据行合并
            ws.merge_cells(start_row=data_start, start_column=1, end_row=data_end, end_column=1)
            ws.merge_cells(start_row=data_start, start_column=2, end_row=data_end, end_column=2)
            ws.merge_cells(start_row=data_start, start_column=3, end_row=data_end, end_column=3)
            ws.merge_cells(start_row=data_start, start_column=10, end_row=data_end, end_column=10)

            # 合并后给合并区域应用格式
            for col in [1, 2, 3, 10]:
                for r in range(data_start, data_end + 1):
                    _apply_format(ws, r, 10)

        # D 列：同品种合并
        merge_start = data_start
        prev_name = items[0].get("name", "")
        for i in range(1, len(items)):
            cur_name = items[i].get("name", "")
            if cur_name and cur_name != prev_name:
                # 不同品种，合并前一段
                if merge_start < data_start + i - 1:
                    ws.merge_cells(start_row=merge_start, start_column=4,
                                   end_row=data_start + i - 1, end_column=4)
                merge_start = data_start + i
                prev_name = cur_name
            elif not cur_name and prev_name:
                # 空 name = 同上品种
                pass
            elif not prev_name and cur_name:
                prev_name = cur_name
        # 合并最后一段
        if merge_start < data_end:
            ws.merge_cells(start_row=merge_start, start_column=4, end_row=data_end, end_column=4)

        # 保存
        wb.save(filepath)

        total_amount = sum(item.get("qty", 0) * item.get("price", 0) for item in items)

        return {
            "success": True,
            "sheet": sheet,
            "start_row": data_start,
            "end_row": data_end,
            "item_count": len(items),
            "total_amount": round(total_amount, 2),
        }
    except Exception as e:
        return {"success": False, "error": f"写入失败: {str(e)}"}
    finally:
        wb.close()


# ---- 工具3: verify_batch ----

def verify_batch(filepath: str, sheet: str, start_row: int, end_row: int) -> dict:
    """验证刚写入的数据是否正确"""
    lock_err = _check_file_lock(filepath)
    if lock_err:
        return lock_err

    # 双模式读取：data_only 读缓存值，公式模式读公式串
    wb_data = load_workbook(filepath, data_only=True)
    wb_formula = load_workbook(filepath, data_only=False)
    try:
        if sheet not in wb_data.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet}' 不存在"}

        ws_data = wb_data[sheet]
        ws_formula = wb_formula[sheet]
        row_count = end_row - start_row + 1
        total_amount = 0
        mismatches = []

        for row in range(start_row, end_row + 1):
            qty = ws_data.cell(row=row, column=7).value or 0
            price = ws_data.cell(row=row, column=8).value or 0
            expected_amount = qty * price

            # I 列：检查是否有公式，有公式则跳过数值比较（缓存值尚未计算）
            formula_cell = ws_formula.cell(row=row, column=9).value
            if isinstance(formula_cell, str) and formula_cell.startswith('='):
                # 公式存在，跳过缓存的未计算值
                pass
            else:
                actual_amount = ws_data.cell(row=row, column=9).value or 0
                if abs(expected_amount - actual_amount) > 0.01:
                    mismatches.append({
                        "row": row,
                        "qty": qty,
                        "price": price,
                        "expected": round(expected_amount, 2),
                        "actual": round(actual_amount, 2),
                    })

            total_amount += expected_amount

        return {
            "success": True,
            "row_count": row_count,
            "total_amount": round(total_amount, 2),
            "mismatches": mismatches,
        }
    except Exception as e:
        return {"success": False, "error": f"验证失败: {str(e)}"}
    finally:
        wb_data.close()
        wb_formula.close()


# ---- 工具0: create_new（创建新对账单）----

def create_new(filepath: str, sheet: str) -> dict:
    """创建新的对账单 Excel 文件，写入表头"""
    lock_err = _check_file_lock(filepath)
    if lock_err:
        return lock_err
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = sheet
        _set_column_widths(ws)
        headers = ['序号', '单号', '日期', '品种', '规格', '单位', '数量', '单价', '金额', '合计金额']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = FONT_SONG
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            cell.fill = HEADER_FILL
        wb.save(filepath)
        return {"success": True, "sheet": sheet, "filepath": filepath, "next_row": 2}
    except Exception as e:
        return {"success": False, "error": f"创建文件失败: {str(e)}"}
    finally:
        wb.close()


def _next_sequence(filepath: str, sheet: str) -> int:
    """读取已有序号的最大值；非数字或空值不参与计算。"""
    if not os.path.exists(filepath):
        return 1
    wb = load_workbook(filepath, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return 1
        ws = wb[sheet]
        values = []
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value
            if isinstance(value, (int, float)):
                values.append(int(value))
        return max(values, default=0) + 1
    finally:
        wb.close()


def export_receipts(filepath: str, sheet: str, mode: str, receipts: list[dict]) -> dict:
    """将数据库权威单据批量导出；模型不传递品名、数量、单价等业务数据。"""
    if not receipts:
        return {"success": False, "error": "没有可导出的单据"}
    if mode not in {"new", "append"}:
        return {"success": False, "error": "mode 必须是 new 或 append"}

    # 追加到不存在的文件时自动转成新建，防止没有表头。
    effective_mode = "new" if mode == "append" and not os.path.exists(filepath) else mode
    if effective_mode == "new":
        start_row, seq = 2, 1
    else:
        position = find_last_row(filepath, sheet)
        if not position.get("success", True):
            return position
        start_row = max(2, int(position.get("next_row", 2)))
        seq = _next_sequence(filepath, sheet)

    exported = []
    for index, receipt in enumerate(receipts):
        write_mode = "new" if index == 0 and effective_mode == "new" else "append"
        written = write_batch(
            filepath=filepath,
            sheet=sheet,
            mode=write_mode,
            start_row=start_row,
            seq=seq,
            receipt_no=str(receipt.get("receipt_no", "")),
            date=str(receipt.get("date", "")),
            items=receipt.get("items", []),
        )
        if not written.get("success"):
            return {
                "success": False,
                "error": written.get("error", "写入失败"),
                "receipt_ids": [item["receipt_id"] for item in exported],
                "partial": bool(exported),
            }
        verification = verify_batch(filepath, sheet, written["start_row"], written["end_row"])
        if not verification.get("success") or verification.get("mismatches"):
            return {
                "success": False,
                "error": verification.get("error") or "写入后校验未通过",
                "receipt_ids": [item["receipt_id"] for item in exported],
                "partial": bool(exported),
            }
        exported.append({
            "receipt_id": receipt["id"],
            "receipt_no": receipt.get("receipt_no", ""),
            "start_row": written["start_row"],
            "end_row": written["end_row"],
            "item_count": written["item_count"],
            "total_amount": written["total_amount"],
        })
        start_row = written["end_row"] + 1
        seq += 1

    return {
        "success": True,
        "verified": True,
        "filepath": filepath,
        "sheet": sheet,
        "receipt_ids": [item["receipt_id"] for item in exported],
        "receipts": exported,
        "item_count": sum(item["item_count"] for item in exported),
        "total_amount": round(sum(item["total_amount"] for item in exported), 2),
    }

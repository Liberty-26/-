"""
SteelDigitize Pro — 夸克扫描王识别封装（方案3：image-to-excel → xlsx 内存解析）

识别引擎：夸克扫描王开放平台
  - 场景 image-to-excel：图片直接还原为按行列对齐的表格（xlsx）
  - 该场景为 Agent 专用能力，必须走 Agent 通道：统一调用官方 CLI（yescan），
    由 yescan 以标准 X-Appbuilder-From=cli 与 SCAN_WEBSERVICE_KEY 完成鉴权，
    不直接发 REST 请求（REST 通道会返回 A0102）。
  - 本模块把 xlsx 在内存中解析为结构化行（不落盘、不生成交付表格）
  - 只取 名称及规格/单位/数量/单价；金额列不读取（每行金额由前端计算）
  - 单号/日期从表头文本行提取（不裁剪，表格外信息保留）
名称/规格拆分、品名库归一化由后续 AI 审核（纯代码校准）完成。
"""
from __future__ import annotations

import re
import os
import json
import asyncio
import shutil
import tempfile
import base64
from io import BytesIO
from datetime import date as _date

from openpyxl import load_workbook
import config

# 全角数字/字母 → 半角
_FW = str.maketrans(
    "０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
    "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ．",
    "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.",
)


def _call_yescan_lib(img_bytes: bytes, api_key: str, scene: str) -> dict:
    """同步调用 yescan SDK（Agent 通道，X-Appbuilder-From=cli），返回 xlsx 字节。
    放在线程池执行，避免阻塞事件循环。
    """
    try:
        from yescan.ocr_client import QuarkOCRClient
    except ImportError:
        return {"success": False, "error": "识别组件缺失，请重新安装 yescan"}

    work_dir = tempfile.mkdtemp(prefix="yescan_work_")
    img_path = os.path.join(work_dir, "input.jpg")
    try:
        with open(img_path, "wb") as f:
            f.write(img_bytes)
        with QuarkOCRClient(api_key=api_key, scene=scene, data_type="image") as client:
            result = client.recognize(image_path=img_path)
        if result.code != "00000":
            return {"success": False, "error": f"{result.code} — {result.message or '识别失败'}"}
        data = result.data or {}
        try:
            xlsx_b64 = data["TypesetInfo"][0]["FileBase64"]
        except (KeyError, IndexError, TypeError):
            return {"success": False, "error": "识别返回格式异常，未找到表格数据"}
        try:
            return {"success": True, "raw": base64.b64decode(xlsx_b64)}
        except Exception:
            return {"success": False, "error": "识别结果解码失败"}
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


async def call_quark_excel(image_base64: str, api_key: str | None = None) -> dict:
    """调用夸克 Agent 通道（yescan SDK）执行 image-to-excel，返回 xlsx 字节。

    并发触发夸克 QPS 限流（A0300）时自动退避重试（最多 3 次）。

    Returns:
        {"success": True, "raw": <xlsx bytes>}
        或 {"success": False, "error": "..."}
    """
    api_key = api_key or config.SCAN_API_KEY
    if not api_key:
        return {"success": False, "error": "识别 API Key 未配置"}

    clean_b64 = image_base64.strip()
    if clean_b64.startswith("data:"):
        if ";base64," in clean_b64:
            clean_b64 = clean_b64.split(";base64,", 1)[1]
        else:
            return {"success": False, "error": "base64 数据格式错误"}

    try:
        img_bytes = base64.b64decode(clean_b64)
    except Exception:
        return {"success": False, "error": "图片 base64 解码失败"}
    if not img_bytes:
        return {"success": False, "error": "图片内容为空"}

    last_error = ""
    for attempt in range(3):
        result = await asyncio.to_thread(_call_yescan_lib, img_bytes, api_key, config.SCAN_SCENE)
        if result.get("success"):
            return result
        last_error = result.get("error", "")
        if "A0300" in last_error and attempt < 2:
            await asyncio.sleep(1.5 * (attempt + 1))
            continue
        return result
    return {"success": False, "error": last_error}


def _strip(s) -> str:
    """单元格清洗：去勾选符号、全角转半角"""
    if s is None:
        return ""
    s = str(s).replace("✔", "").replace("✓", "").replace("√", "")
    return s.translate(_FW).strip()


def _to_float(s) -> float:
    """数量/单价安全转 float：容忍 '10元'/'015.75'/全角"""
    if s is None:
        return 0.0
    text = _strip(s).replace("元", "").replace("¥", "").replace("￥", "").strip()
    m = re.search(r"\d+(?:\.\d+)?", text)
    return float(m.group(0)) if m else 0.0


def _extract_receipt_no(text: str) -> str:
    m = re.search(r"送\s*货\s*单\s*[:：]?\s*([A-Za-z0-9\-]{3,})", text)
    if m:
        return m.group(1).strip()
    m = re.search(r"单号\s*[:：]?\s*([A-Za-z0-9\-]{3,})", text)
    return m.group(1).strip() if m else ""


def _extract_date(text: str) -> tuple[str, bool]:
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
    if not m:
        m = re.search(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", text)
    if not m:
        return "", False
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    date_str = f"{y:04d}-{mo:02d}-{d:02d}"
    return date_str, abs(y - _date.today().year) > 5


def parse_receipt(xlsx_bytes: bytes) -> dict:
    """xlsx → 单据字段 + 明细行（金额列不读取，前端按 数量×单价 计算）"""
    try:
        ws = load_workbook(BytesIO(xlsx_bytes), read_only=True).active
        rows = [[_strip(c) for c in row] for row in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(r)]
    except Exception as e:
        return {"success": False, "error": f"xlsx 解析失败: {e}"}

    head_text = "\n".join(" ".join(r) for r in rows[:12])
    receipt_no = _extract_receipt_no(head_text)
    date_str, suspicious = _extract_date(head_text)

    # 表头定位（序号/名称及规格/单位/数量/单价/金额/备注）
    hdr_idx = next((i for i, r in enumerate(rows)
                    if any("名称" in c and "规格" in c for c in r)), None)
    if hdr_idx is None:
        return {"success": True, "receipt_no": receipt_no, "date": date_str,
                "date_suspicious": suspicious, "items": [], "raw_response": head_text}

    hdr = rows[hdr_idx]
    col = {name: i for i, name in enumerate(hdr) if name}
    idx_name = next((v for k, v in col.items() if "名称" in k), None)
    idx_unit = col.get("单位")
    idx_qty = col.get("数量")
    idx_price = col.get("单价")
    idx_amount = col.get("金额")

    items = []
    total_candidate = None   # 小计兜底
    rec_total = None         # 合计/总计（最终总金额）
    for r in rows[hdr_idx + 1:]:
        first = (r[0] or "") if r else ""
        if any(k in first for k in ("合计", "小计", "总计", "大写")):
            # 从金额列（或行内最后一个数字）提取合计金额
            raw_amt = ""
            if idx_amount is not None and idx_amount < len(r):
                raw_amt = _strip(r[idx_amount])
            if not raw_amt:
                for cell in reversed(r[1:]):
                    if _strip(cell):
                        raw_amt = _strip(cell)
                        break
            val = _to_float(raw_amt) if raw_amt else None
            if val is not None and val > 0:
                if "合计" in first or "总计" in first:
                    rec_total = val
                    break
                total_candidate = val  # 小计：先兜底，继续找合计
            if "合计" in first or "总计" in first:
                break
            continue
        if idx_name is None or idx_name >= len(r):
            continue
        name = r[idx_name] if idx_name < len(r) else ""
        if not name:
            continue
        unit = _strip(r[idx_unit]) if idx_unit is not None and idx_unit < len(r) else ""
        qty = _to_float(r[idx_qty]) if idx_qty is not None and idx_qty < len(r) else 0.0
        price = _to_float(r[idx_price]) if idx_price is not None and idx_price < len(r) else 0.0
        raw_amt = _strip(r[idx_amount]) if idx_amount is not None and idx_amount < len(r) else ""
        rec_amount = _to_float(raw_amt) if raw_amt else None
        item = {"name": name, "spec": "", "unit": unit, "qty": qty, "price": price}
        if rec_amount is not None:
            item["rec_amount"] = rec_amount
        items.append(item)

    if rec_total is None:
        rec_total = total_candidate
    return {"success": True, "receipt_no": receipt_no, "date": date_str,
            "date_suspicious": suspicious, "items": items,
            "rec_total": rec_total, "raw_response": head_text}
